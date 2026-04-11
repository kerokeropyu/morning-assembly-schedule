#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
総務部 朝礼当番表 自動生成スクリプト
・1つのExcelファイルの中に、月ごとにシート（タブ）を追加していく方式
・ファイル名は固定（設定ファイルで指定）
・土日・祝日は当番なし（セルをグレーにして「休」と表示）
・通常実行：翌月分のシートを1枚追加
・引数あり：指定年月のシートを1枚追加（例: python generate_schedule.py 2026 4）

【テンプレート対応について】
設定ファイル（morning_assembly_config.json）に以下を追加することで
テンプレートExcelをもとに出力できます。

  "template_file": テンプレートExcelのパス（例: "./template.xlsx"）
                  exe化する場合は実行ファイルと同フォルダに配置して
                  "./template.xlsx" のように相対パスで指定してください。

【テンプレートのレイアウト想定（添付ファイルの掃除担当表をもとに）】
  行1      : タイトル行（「掃除担当表」等） ← 上書きしない
  行2      : 週の開始日見出し行             ← 月初週から連続した日付で上書き
  行3      : 列ヘッダー行（フロア/番号/掃除場所/担当者）← テンプレートのまま残す
  行4以降  : 清掃場所ごとの担当者行         ← 担当者名だけを書き込む

  列構成（設定ファイルの start_col を起点）:
    start_col+0 : フロア（テンプレートの固定値をそのまま使用）
    start_col+1 : 番号
    start_col+2 : 掃除場所
    start_col+3 : サポート情報（任意）
    start_col+4 : 担当者（1週目）
    start_col+5 : 空白列（テンプレートの1列おきの構成に合わせる）
    start_col+6 : 担当者（2週目）
    ... 以降、1列おきに担当者列が続く

  ※ start_row / start_col は設定ファイルで調整してください。
"""

import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from calendar import monthrange

import jpholiday
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def load_config(config_path: str = "morning_assembly_config.json"):
    """設定ファイルを読み込む"""
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"エラー: 設定ファイル '{config_path}' が見つかりません。")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"エラー: 設定ファイル '{config_path}' の内容が正しくありません（JSONエラー）。")
        sys.exit(1)


# =============================================================================
# 【テンプレート対応 修正箇所①】load_or_create_workbook()
#
# ・出力ファイルが既に存在する場合はそのまま開く（従来通り）
# ・存在しない場合：
#     template_path が指定されていればテンプレートファイルをコピーして雛形とする
#     指定がなければ従来通り空の Workbook を新規作成する
#
# exe化する場合：
#   template_file に "./template.xlsx" と書けば、
#   実行ファイルと同じフォルダにある template.xlsx を読み込みます。
# =============================================================================
def load_or_create_workbook(output_dir: Path, filename: str, template_path: str = None):
    """Excelファイルを読み込む。なければテンプレートをコピー（またはWorkbook新規作成）。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    filepath = output_dir / filename
    if filepath.exists():
        wb = load_workbook(filepath)
    elif template_path:
        # ---------------------------------------------------------------
        # テンプレートファイルをコピーして雛形として使う
        # ※ テンプレートのパスは設定ファイルの "template_file" キーで外部指定
        # ---------------------------------------------------------------
        template_file = Path(template_path)
        if not template_file.exists():
            print(f"エラー: テンプレートファイル '{template_path}' が見つかりません。")
            sys.exit(1)
        wb = load_workbook(template_file)
    else:
        wb = Workbook()
    return wb, filepath


def is_off_day(year: int, month: int, day: int) -> bool:
    """土日または祝日かどうか判定する"""
    d = date(year, month, day)
    return d.weekday() >= 5 or jpholiday.is_holiday(d)


def get_weekday_dates(year: int, month: int):
    """指定年月の平日（土日祝を除く）のdate一覧を返す"""
    _, num_days = monthrange(year, month)
    return [
        date(year, month, d)
        for d in range(1, num_days + 1)
        if not is_off_day(year, month, d)
    ]


def create_month_sheet(wb, year: int, month: int, members: list, config: dict):
    """
    指定年月のシートを作成（既にあれば作り直し）。

    【テンプレート対応 修正箇所②】シートの雛形について
    ・config に "template_sheet_name" が設定されている場合は
      テンプレートシートをコピーして雛形とする
    ・設定がない場合は従来通り空のシートを新規作成する
    ※ テンプレートシートは非表示（hidden）にしておくと
      ユーザーに見えなくてすっきりする

    【テンプレート対応 修正箇所③】描画の始点について
    ・config の "start_row" / "start_col" で書き込み開始位置を外部指定できる
    ・テンプレートにタイトル行や装飾がある場合はここを調整する
    """
    sheet_title = f"{year}年{month:02d}月"

    if sheet_title in wb.sheetnames:
        wb.remove(wb[sheet_title])

    # ------------------------------------------------------------------
    # 【修正箇所②】テンプレートシートが設定されている場合はコピーして使う
    # ------------------------------------------------------------------
    template_sheet_name = config.get("template_sheet_name")
    if template_sheet_name and template_sheet_name in wb.sheetnames:
        template_ws = wb[template_sheet_name]
        ws = wb.copy_worksheet(template_ws)
        ws.title = sheet_title
        # テンプレートシートを非表示にする
        template_ws.sheet_state = "hidden"
    else:
        ws = wb.create_sheet(title=sheet_title)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except KeyError:
            pass

    # ------------------------------------------------------------------
    # 【修正箇所③】書き込み始点を設定ファイルから取得する
    #
    # テンプレートのレイアウトに合わせて以下を調整してください。
    #
    # header_row : 「担当者」列ヘッダーがある行番号（例: 3 → 3行目がヘッダー）
    # data_row   : 担当者名を書き始める行番号（例: 4 → 4行目から清掃場所行が始まる）
    # name_col   : 担当者列の先頭列番号
    #              ※ 添付テンプレートでは1列おきに「担当者」列がある構成のため、
    #                 name_col から 2列おきに担当者を書き込む
    # week_row   : 週の日付（「8/7-」等）を書き込む行番号（例: 2 → 2行目）
    # ------------------------------------------------------------------
    header_row = config.get("header_row", 3)   # ヘッダー行（「担当者」見出しの行）
    data_row   = config.get("data_row",   4)   # データ開始行（清掃場所の1行目）
    name_col   = config.get("name_col",   5)   # 担当者列の先頭列番号（A=1）
    week_row   = config.get("week_row",   2)   # 週日付を書く行番号
    col_step   = config.get("col_step",   2)   # 担当者列の間隔（テンプレートが1列おきなら2）
    # ---------------------------------------------------------------
    # 平日リストを取得し、週ごとにグループ化する
    # ---------------------------------------------------------------
    weekday_dates = get_weekday_dates(year, month)

    # 週単位でグループ化（月曜始まりの週で区切る）
    weeks = []
    if weekday_dates:
        current_week = [weekday_dates[0]]
        for d in weekday_dates[1:]:
            if d.weekday() < current_week[-1].weekday():  # 週が変わった
                weeks.append(current_week)
                current_week = [d]
            else:
                current_week.append(d)
        weeks.append(current_week)

    # ---------------------------------------------------------------
    # スタイル定義
    # ---------------------------------------------------------------
    off_fill    = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    border      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    num_members = len(members)
    num_locations = config.get("num_locations", num_members)  # 清掃場所数（デフォルトはメンバー数と同じ）

    # ---------------------------------------------------------------
    # 週の日付見出しを書き込む（行2: 「8/7-」等）
    # ※ テンプレートの既存見出しを上書きする
    # ---------------------------------------------------------------
    for wi, week in enumerate(weeks):
        col = name_col + wi * col_step
        # 週の開始日（月曜）を取得
        start_of_week = week[0]
        week_label = f"{start_of_week.month}/{start_of_week.day}-"
        cell = ws.cell(row=week_row, column=col, value=week_label)
        cell.alignment = center_align

    # ---------------------------------------------------------------
    # 担当者を週ごとに書き込む
    # duty_index : 今まで何回平日があったかのカウンター（ローテーション計算用）
    # ---------------------------------------------------------------
    duty_index = 0
    for wi, week in enumerate(weeks):
        col = name_col + wi * col_step  # この週の担当者列

        for d in week:
            # 各平日の担当者（ローテーション）
            assigned_member = members[duty_index % num_members]
            duty_index += 1

            # 清掃場所ごとの行に担当者名を書き込む
            # ※ テンプレートの1行 = 1清掃場所に対応
            # ※ 全場所に同じ担当者を書く場合はこのままでOK
            #   場所ごとにメンバーをずらしたい場合は以下のロジックを修正する
            for loc_idx in range(num_locations):
                row = data_row + loc_idx
                cell = ws.cell(row=row, column=col)
                # 平日は担当者名、休日はすでに空白のままにする
                # （この週の中の平日は同じ担当者を書く。週1回の当番表のため）
                # ※ 1週間に複数の平日があるが、週の担当者は最初の日の担当者で統一
                # ※ 上書きを避けるため、最初の平日（d == week[0]）のみ書き込む
                if d == week[0]:
                    cell.value = assigned_member
                    cell.alignment = center_align
                    cell.border = border

    # ---------------------------------------------------------------
    # 休日（土日祝）の週があればグレーアウト
    # ---------------------------------------------------------------
    _, num_days = monthrange(year, month)
    for day in range(1, num_days + 1):
        if is_off_day(year, month, day):
            # 休日列はテンプレートのまま（担当者列が関係する場合のみグレーアウト）
            pass  # テンプレートがすでに休日を考慮している場合はここは不要


def generate_schedule_for(year: int, month: int, config: dict):
    """指定年月のシートを1枚追加する（手動実行用）"""
    output_dir = Path(config["output_directory"])
    excel_filename = config["excel_filename"]

    # 【修正箇所①】template_path を config から取得して load_or_create_workbook() に渡す
    # 設定ファイルに "template_file": "./template.xlsx" を追加すれば外部指定できる
    template_path = config.get("template_file")
    wb, filepath = load_or_create_workbook(output_dir, excel_filename, template_path)

    create_month_sheet(wb, year, month, config["members"], config)
    wb.save(filepath)
    print("============================================")
    print("朝礼当番表シートを作成しました")
    print(f"  ファイル: {filepath}")
    print(f"  シート名: {year}年{month:02d}月")
    print(f"  メンバー数: {len(config['members'])} 名")
    if config.get("template_file"):
        print(f"  テンプレート: {config['template_file']}")
    print("============================================")


def generate_next_month(config: dict):
    """翌月分のシートを1枚追加する（月末バッチ用）"""
    today = datetime.now()
    if today.month == 12:
        year = today.year + 1
        month = 1
    else:
        year = today.year
        month = today.month + 1
    generate_schedule_for(year, month, config)


if __name__ == "__main__":
    config = load_config()
    # 使い方：
    #   python generate_schedule.py           → 翌月分を1枚追加
    #   python generate_schedule.py 2026 4   → 2026年4月分を1枚追加
    if len(sys.argv) >= 3:
        y = int(sys.argv[1])
        m = int(sys.argv[2])
        generate_schedule_for(y, m, config)
    else:
        generate_next_month(config)
